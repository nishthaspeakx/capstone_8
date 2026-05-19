"""
Build a comprehensive feature documentation Excel workbook.
Outputs: outputs/Phase1_Feature_Documentation.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ──────────────────────────────────────────────────────────────
# DATA: All 35 features with full documentation
# ──────────────────────────────────────────────────────────────
FEATURES = [
    # ═══ CALENDAR (2) ═══
    {
        "no": 1, "group": "A. Calendar", "name": "Year",
        "type": "Raw", "dtype": "Integer", "source": "Direct from CSV",
        "formula": "df['Year']",
        "corr_target": -0.059, "importance_pct": 0.30,
        "business_logic": "Captures macro-level chain-wide drift over time. Lowe's chain saw average store sales decline ~13% from 2023 ($1.25M/wk) to 2025 ($1.08M/wk).",
        "why_included": "Without Year, the model would treat 2023 and 2025 weeks as identical when in fact macro conditions changed (inflation, housing market shifts, post-COVID normalization).",
        "leakage_safe": "YES — fully known at prediction time",
        "data_quality": "0% missing, 3 unique values (2023, 2024, 2025)",
        "example_value": "2025",
        "alternative_considered": "Could have used 'years since 2023' but Year is more interpretable for stakeholders"
    },
    {
        "no": 2, "group": "A. Calendar", "name": "Fiscal Week",
        "type": "Raw", "dtype": "Integer", "source": "Direct from CSV",
        "formula": "df['Fiscal Week']",
        "corr_target": -0.086, "importance_pct": 0.50,
        "business_logic": "Captures weekly seasonality. Spring weeks (gardening season) peak at $1.47M/wk while winter weeks dip to $753K — a 96% peak-to-trough spread.",
        "why_included": "Home improvement is HIGHLY seasonal. Without Fiscal Week, the model can't differentiate a slow February from a peak May. Critical for monthly target allocation.",
        "leakage_safe": "YES — fully known at prediction time",
        "data_quality": "0% missing, 52 unique values (1-52)",
        "example_value": "33 (mid-August week)",
        "alternative_considered": "Could use month or quarter, but week-level granularity is essential for retail planning"
    },

    # ═══ ENGINEERED LAGS (6) — THE POWER FEATURES ═══
    {
        "no": 3, "group": "B. Engineered Lags", "name": "lag_1",
        "type": "Engineered", "dtype": "Float (USD)",
        "source": "Engineered from Actual Sales USD",
        "formula": "df.groupby('Store ID')['Actual Sales USD'].shift(1)",
        "corr_target": 0.982, "importance_pct": 49.86,
        "business_logic": "Last week's sales for the SAME store. By far the strongest predictor — sales have strong week-to-week persistence.",
        "why_included": "STRONGEST single predictor in the entire dataset (corr=0.982). Captures very recent store-specific momentum.",
        "leakage_safe": "YES — uses shift(1) so prediction at week t only sees data up to week t-1",
        "data_quality": "First week per store = NaN (no history); ~1,727 NaN rows dropped",
        "example_value": "$908,622",
        "alternative_considered": "Could use only this lag, but adding lag_4/13/52 captures longer rhythms"
    },
    {
        "no": 4, "group": "B. Engineered Lags", "name": "lag_4",
        "type": "Engineered", "dtype": "Float (USD)",
        "source": "Engineered from Actual Sales USD",
        "formula": "df.groupby('Store ID')['Actual Sales USD'].shift(4)",
        "corr_target": 0.964, "importance_pct": 0.10,
        "business_logic": "Sales from 4 weeks ago = approximately one month ago. Captures monthly rhythm (pay cycles, monthly promotions, billing).",
        "why_included": "Many home improvement decisions follow monthly cycles. Bridge between weekly and quarterly patterns.",
        "leakage_safe": "YES — shift(4) ensures only past data used",
        "data_quality": "First 4 weeks per store = NaN",
        "example_value": "$844,976",
        "alternative_considered": "lag_2 or lag_8 also viable — chose 4 for monthly alignment"
    },
    {
        "no": 5, "group": "B. Engineered Lags", "name": "lag_13",
        "type": "Engineered", "dtype": "Float (USD)",
        "source": "Engineered from Actual Sales USD",
        "formula": "df.groupby('Store ID')['Actual Sales USD'].shift(13)",
        "corr_target": 0.913, "importance_pct": 0.05,
        "business_logic": "Sales from 13 weeks ago = approximately one quarter ago. Captures quarterly fiscal patterns.",
        "why_included": "Quarter-over-quarter shifts (Q1 vs Q2, etc.) — important for retail seasonality.",
        "leakage_safe": "YES — shift(13)",
        "data_quality": "First 13 weeks per store = NaN",
        "example_value": "$1,118,901",
        "alternative_considered": "lag_12 would give similar monthly read; 13 aligns with retail quarterly definition"
    },
    {
        "no": 6, "group": "B. Engineered Lags", "name": "lag_52",
        "type": "Engineered", "dtype": "Float (USD)",
        "source": "Engineered from Actual Sales USD",
        "formula": "df.groupby('Store ID')['Actual Sales USD'].shift(52)",
        "corr_target": 0.953, "importance_pct": 0.98,
        "business_logic": "Sales from 52 weeks ago = same week LAST YEAR. Captures pure annual seasonality.",
        "why_included": "Memorial Day 2024 → predicts Memorial Day 2025. The most reliable seasonal anchor for any retail business.",
        "leakage_safe": "YES — shift(52)",
        "data_quality": "First year per store = NaN (only 2024+ data has lag_52)",
        "example_value": "$861,899",
        "alternative_considered": "This is also our naive baseline (8.27% WAPE) — model must beat this to be useful"
    },
    {
        "no": 7, "group": "B. Engineered Lags", "name": "roll_4",
        "type": "Engineered", "dtype": "Float (USD)",
        "source": "Engineered from Actual Sales USD",
        "formula": "groupby('Store ID')['Actual Sales USD'].transform(lambda x: x.shift(1).rolling(4).mean())",
        "corr_target": 0.981, "importance_pct": 43.76,
        "business_logic": "Average of the last 4 weeks of sales (excluding current week). Smooths out one-week spikes/dips to reveal trend.",
        "why_included": "SECOND most important feature. While lag_1 captures the latest point, roll_4 captures the recent trend trajectory.",
        "leakage_safe": "YES — shift(1) BEFORE rolling ensures no current-week leak",
        "data_quality": "First 5 weeks per store = NaN (4 for window + 1 for shift)",
        "example_value": "$951,491",
        "alternative_considered": "Critical to apply shift(1) BEFORE rolling — otherwise current week leaks in"
    },
    {
        "no": 8, "group": "B. Engineered Lags", "name": "roll_13",
        "type": "Engineered", "dtype": "Float (USD)",
        "source": "Engineered from Actual Sales USD",
        "formula": "groupby('Store ID')['Actual Sales USD'].transform(lambda x: x.shift(1).rolling(13).mean())",
        "corr_target": 0.964, "importance_pct": 0.37,
        "business_logic": "Average of the last 13 weeks (one quarter) of sales. Smooths out monthly noise to reveal quarterly trend.",
        "why_included": "Captures the slower-moving seasonal trend without short-term volatility (storms, one-off promos).",
        "leakage_safe": "YES — shift(1) before rolling",
        "data_quality": "First 14 weeks per store = NaN",
        "example_value": "$973,518",
        "alternative_considered": "roll_8 (half-quarter) and roll_26 (half-year) tested but added little"
    },

    # ═══ STORE/MARKET (5) ═══
    {
        "no": 9, "group": "C. Store/Market", "name": "Sales Floor Size",
        "type": "Raw", "dtype": "Float (sq ft)",
        "source": "Direct from CSV",
        "formula": "df['Sales Floor Size']",
        "corr_target": 0.035, "importance_pct": 0.13,
        "business_logic": "Square footage of indoor selling space. Bigger floor = more product assortment = higher sales potential.",
        "why_included": "Acts as a NORMALIZER for other features. A small store with $1M lag_1 is overperforming; a large store with same lag_1 is underperforming.",
        "leakage_safe": "YES — physical attribute, doesn't change weekly",
        "data_quality": "0% missing",
        "example_value": "114,111 sq ft",
        "alternative_considered": "PDF noted fulfillment is largely size-independent, but it still matters for walk-in experience"
    },
    {
        "no": 10, "group": "C. Store/Market", "name": "Garden Ctr Size",
        "type": "Raw", "dtype": "Float (sq ft)",
        "source": "Direct from CSV",
        "formula": "df['Garden Ctr Size']",
        "corr_target": 0.062, "importance_pct": 0.16,
        "business_logic": "Square footage of outdoor garden center. Garden is a 'traffic-driving category' — pulls in seasonal shoppers who buy other products.",
        "why_included": "Explicitly called out as PRIORITY HIGH in the project PDF. Stores with bigger garden centers see seasonal traffic spikes.",
        "leakage_safe": "YES — physical attribute",
        "data_quality": "0% missing, 468 zeros are valid (stores without garden centers)",
        "example_value": "38,520 sq ft",
        "alternative_considered": "Considered a binary 'has garden center' flag but the size variation matters"
    },
    {
        "no": 11, "group": "C. Store/Market", "name": "Urbanicity_enc",
        "type": "Encoded", "dtype": "Integer (0-6)",
        "source": "LabelEncoded from Urbanicity",
        "formula": "LabelEncoder().fit_transform(df['Urbanicity'])",
        "corr_target": "Categorical", "importance_pct": 0.10,
        "business_logic": "7 categories: Metropolis, Large City, Small City, Town, Suburban, Rural, Remote. Captures customer density and lifestyle.",
        "why_included": "MASTER segmentation variable. Metropolis stores avg $1.63M/wk vs Large City at $930K — 75% gap from this single feature.",
        "leakage_safe": "YES — store attribute",
        "data_quality": "0% missing, 7 unique levels",
        "example_value": "4 (Small City)",
        "alternative_considered": "One-hot encoding considered but tree models handle integer encoding well"
    },
    {
        "no": 12, "group": "C. Store/Market", "name": "CBSA Type_enc",
        "type": "Encoded", "dtype": "Integer (0-2)",
        "source": "LabelEncoded from CBSA Type",
        "formula": "LabelEncoder().fit_transform(df['CBSA Type'])",
        "corr_target": "Categorical", "importance_pct": 0.04,
        "business_logic": "Core-Based Statistical Area type: Metro, Micro, or Non-CBSA. Captures regional economic structure.",
        "why_included": "Complements Urbanicity. Urbanicity = local density; CBSA = regional economic integration.",
        "leakage_safe": "YES",
        "data_quality": "0% missing, 3 unique levels",
        "example_value": "2 (Metro)",
        "alternative_considered": "Highly correlated with Urbanicity but adds incremental signal"
    },
    {
        "no": 13, "group": "C. Store/Market", "name": "CBSA Metro Size_enc",
        "type": "Encoded", "dtype": "Integer (0-3)",
        "source": "LabelEncoded from CBSA Metro Size",
        "formula": "LabelEncoder().fit_transform(df['CBSA Metro Size'])",
        "corr_target": "Categorical", "importance_pct": 0.04,
        "business_logic": "Within Metro CBSAs, the size: Small, Medium, Large, Mega. 4 finer-grained categories.",
        "why_included": "Small Metro stores avg $1.27M vs Large Metro $1.06M (counterintuitive — small metros may have less competition).",
        "leakage_safe": "YES",
        "data_quality": "0% missing, 4 unique levels",
        "example_value": "1 (Medium)",
        "alternative_considered": "Could be combined with CBSA Type but adds finer granularity for tree splits"
    },

    # ═══ DEMAND/HOUSING — RAW (11) ═══
    {
        "no": 14, "group": "D. Demand/Housing (Raw)", "name": "CYE Total Households",
        "type": "Raw", "dtype": "Integer",
        "source": "Direct from CSV (Current Year Estimate)",
        "formula": "df['CYE Total Households']",
        "corr_target": -0.045, "importance_pct": 0.10,
        "business_logic": "Number of households in the store's trade area. Total addressable market — more households = more potential customers.",
        "why_included": "Most fundamental demand driver. Std dev = 143K — significant variation across stores.",
        "leakage_safe": "YES — demographic snapshot",
        "data_quality": "0% missing",
        "example_value": "33,452 households",
        "alternative_considered": "Negative correlation because high-density areas have more competition diluting per-store share"
    },
    {
        "no": 15, "group": "D. Demand/Housing (Raw)", "name": "CYE Household Density HH SqMi",
        "type": "Raw", "dtype": "Float",
        "source": "Direct from CSV",
        "formula": "df['CYE Household Density HH SqMi']",
        "corr_target": -0.001, "importance_pct": 0.04,
        "business_logic": "Households per square mile. Density complement to total count.",
        "why_included": "100K households in 5 sq mi (high density, walkable) is very different from 100K in 500 sq mi (low density, car-based).",
        "leakage_safe": "YES",
        "data_quality": "0% missing",
        "example_value": "127.5",
        "alternative_considered": "Near-zero correlation but adds context for tree splits with other features"
    },
    {
        "no": 16, "group": "D. Demand/Housing (Raw)", "name": "CYE Median Household Income",
        "type": "Raw", "dtype": "Float (USD)",
        "source": "Direct from CSV",
        "formula": "df['CYE Median Household Income']",
        "corr_target": -0.036, "importance_pct": 0.10,
        "business_logic": "The MIDDLE household income in trade area. Best single summary of purchasing power.",
        "why_included": "Median is robust to outliers (one billionaire doesn't skew it). PDF marked HIGH priority.",
        "leakage_safe": "YES",
        "data_quality": "0% missing",
        "example_value": "$61,546",
        "alternative_considered": "Mean income considered but median is more representative"
    },
    {
        "no": 17, "group": "D. Demand/Housing (Raw)", "name": "Compound Annual HH Growth Rate 2010 2020",
        "type": "Raw", "dtype": "Float (%)",
        "source": "Direct from CSV",
        "formula": "df['Compound Annual HH Growth Rate 2010 2020']",
        "corr_target": 0.073, "importance_pct": 0.07,
        "business_logic": "Annual % growth in households over 2010-2020 decade. Forward-looking indicator.",
        "why_included": "FORWARD-LOOKING. Used heavily in Role-of-Store segmentation. Identifies growing vs declining markets.",
        "leakage_safe": "YES — historical trend",
        "data_quality": "0% missing, range -1.22% to 6.18%",
        "example_value": "1.85%/year",
        "alternative_considered": "Most positive correlation among static demographics aside from housing_new_share"
    },
    {
        "no": 18, "group": "D. Demand/Housing (Raw)", "name": "CYE Total Housing Units",
        "type": "Raw", "dtype": "Integer",
        "source": "Direct from CSV",
        "formula": "df['CYE Total Housing Units']",
        "corr_target": -0.040, "importance_pct": 0.06,
        "business_logic": "All housing units (occupied + vacant + new construction).",
        "why_included": "Captures renovation demand BEYOND current occupants. Vacant/new units imply construction activity.",
        "leakage_safe": "YES",
        "data_quality": "0% missing",
        "example_value": "35,720 units",
        "alternative_considered": "Correlated with Total Households but vacant/new differential matters"
    },
    {
        "no": 19, "group": "D. Demand/Housing (Raw)", "name": "CYE % Housing Units Owned",
        "type": "Raw", "dtype": "Float (%)",
        "source": "Direct from CSV",
        "formula": "df['CYE % Housing Units Owned']",
        "corr_target": 0.019, "importance_pct": 0.05,
        "business_logic": "Percentage of homes that are owner-occupied (vs. rented).",
        "why_included": "Renters spend 60-70% LESS on home improvement than owners. Fundamental demand driver.",
        "leakage_safe": "YES",
        "data_quality": "0% missing",
        "example_value": "68.5%",
        "alternative_considered": "% Rented dropped (perfect inverse — 100% redundant)"
    },
    {
        "no": 20, "group": "D. Demand/Housing (Raw)", "name": "CYE Median Year Housing Unit Built",
        "type": "Raw", "dtype": "Integer (Year)",
        "source": "Direct from CSV",
        "formula": "df['CYE Median Year Housing Unit Built']",
        "corr_target": 0.078, "importance_pct": 0.07,
        "business_logic": "The MIDDLE year homes were built in trade area.",
        "why_included": "Single best housing-age indicator. Older homes need plumbing/electrical/roofing repairs.",
        "leakage_safe": "YES",
        "data_quality": "0% missing",
        "example_value": "1985",
        "alternative_considered": "Avoids using all 9 decade buckets (multicollinearity)"
    },
    {
        "no": 21, "group": "D. Demand/Housing (Raw)", "name": "CYE Average Mean Length of Residence Years",
        "type": "Raw", "dtype": "Float (Years)",
        "source": "Direct from CSV",
        "formula": "df['CYE Average Mean Length of Residence Years']",
        "corr_target": -0.038, "importance_pct": 0.21,
        "business_logic": "Average years residents have lived in their home.",
        "why_included": "BEHAVIORAL TRIGGER. New move-ins (low) = improvement projects. Long-term residents (high) = maintenance.",
        "leakage_safe": "YES",
        "data_quality": "0% missing",
        "example_value": "12.3 years",
        "alternative_considered": "Top static demographic in feature importance"
    },
    {
        "no": 22, "group": "D. Demand/Housing (Raw)", "name": "CYE Total Population",
        "type": "Raw", "dtype": "Integer",
        "source": "Direct from CSV",
        "formula": "df['CYE Total Population']",
        "corr_target": -0.044, "importance_pct": 0.06,
        "business_logic": "Total people in trade area. Complements household count.",
        "why_included": "Captures multi-person households and individuals. 4-person household has different dynamics than 1-person.",
        "leakage_safe": "YES",
        "data_quality": "0% missing",
        "example_value": "82,510",
        "alternative_considered": "Correlated with Total Households but average household size varies"
    },
    {
        "no": 23, "group": "D. Demand/Housing (Raw)", "name": "CYE Median Age Total Pop",
        "type": "Raw", "dtype": "Float (Years)",
        "source": "Direct from CSV",
        "formula": "df['CYE Median Age Total Pop']",
        "corr_target": 0.034, "importance_pct": 0.04,
        "business_logic": "Middle age of population in trade area.",
        "why_included": "Age 35-55 is PEAK DIY bracket — homeowners with disposable income and energy.",
        "leakage_safe": "YES",
        "data_quality": "0% missing",
        "example_value": "39.2",
        "alternative_considered": "Could engineer 'age 35-55 share' but median captures central tendency"
    },
    {
        "no": 24, "group": "D. Demand/Housing (Raw)", "name": "CYE Veteran Population",
        "type": "Raw", "dtype": "Integer",
        "source": "Direct from CSV",
        "formula": "df['CYE Veteran Population']",
        "corr_target": -0.018, "importance_pct": 0.06,
        "business_logic": "Number of veterans in trade area.",
        "why_included": "LOWE'S-SPECIFIC: 10% permanent veteran discount. Higher veteran pop = more traffic AND lower margins.",
        "leakage_safe": "YES",
        "data_quality": "0% missing",
        "example_value": "5,840",
        "alternative_considered": "Unique business-specific feature; absent from generic demographic models"
    },

    # ═══ DEMAND/HOUSING — ENGINEERED (5) ═══
    {
        "no": 25, "group": "E. Demand/Housing (Engineered)", "name": "income_low_share",
        "type": "Engineered", "dtype": "Float (%)",
        "source": "Engineered from 2 income brackets",
        "formula": "df['CYE % HH Income < 25k'] + df['CYE % HH Income 25k-35k']",
        "corr_target": 0.027, "importance_pct": 0.04,
        "business_logic": "Combined share of households earning <$35K. Captures 'value-conscious' segment.",
        "why_included": "Replaces 2 raw income bins. Captures budget-shopper density without multicollinearity.",
        "leakage_safe": "YES — demographic snapshot",
        "data_quality": "0% missing after engineering",
        "example_value": "18.5%",
        "alternative_considered": "Using all 8 income brackets creates multicollinearity (sum to 100%)"
    },
    {
        "no": 26, "group": "E. Demand/Housing (Engineered)", "name": "income_diy_core",
        "type": "Engineered", "dtype": "Float (%)",
        "source": "Engineered from 2 income brackets",
        "formula": "df['CYE % HH Income 50k-75k'] + df['CYE % HH Income 75k-100K']",
        "corr_target": 0.040, "importance_pct": 0.07,
        "business_logic": "Combined share of $50K-$100K households. The CORE DIY customer segment.",
        "why_included": "These households have enough income for projects but not enough to hire everything out — Lowe's bread-and-butter.",
        "leakage_safe": "YES",
        "data_quality": "0% missing",
        "example_value": "27.3%",
        "alternative_considered": "$50K-$100K is the empirical sweet spot for home improvement spend"
    },
    {
        "no": 27, "group": "E. Demand/Housing (Engineered)", "name": "income_affluent",
        "type": "Engineered", "dtype": "Float (%)",
        "source": "Engineered from 2 income brackets",
        "formula": "df['CYE % HH Income 150k-250k'] + df['CYE % HH Income 250K+']",
        "corr_target": -0.032, "importance_pct": 0.18,
        "business_logic": "Combined share of $150K+ households. Pro/contractor proxy.",
        "why_included": "Wealthier areas → more remodeling, higher-value purchases, contractor-driven volume.",
        "leakage_safe": "YES",
        "data_quality": "0% missing after engineering (CYE % HH Income 250K+ had 41.6% missing — handled by fillna(0))",
        "example_value": "12.7%",
        "alternative_considered": "Negative correlation despite high importance — captures edge segment behavior"
    },
    {
        "no": 28, "group": "E. Demand/Housing (Engineered)", "name": "housing_old_share",
        "type": "Engineered", "dtype": "Float (%)",
        "source": "Engineered from 3 housing decade columns",
        "formula": "df['% Built before 1949'] + df['% Built 1950-1959'] + df['% Built 1960-1969']",
        "corr_target": -0.075, "importance_pct": 0.05,
        "business_logic": "Combined share of pre-1969 housing. Old housing stock = continuous repair demand.",
        "why_included": "Replaces 3 raw decade columns. Old homes need plumbing, electrical, roofing repairs continuously.",
        "leakage_safe": "YES",
        "data_quality": "0% missing after engineering",
        "example_value": "22.4%",
        "alternative_considered": "Using all 9 housing decade buckets creates multicollinearity"
    },
    {
        "no": 29, "group": "E. Demand/Housing (Engineered)", "name": "housing_new_share",
        "type": "Engineered", "dtype": "Float (%)",
        "source": "Engineered from 2 housing decade columns",
        "formula": "df['% Built 2000-2009'] + df['% Built 2009-2019']",
        "corr_target": 0.088, "importance_pct": 0.07,
        "business_logic": "Combined share of post-2000 housing. New construction = new homeowner demand.",
        "why_included": "HIGHEST static-feature correlation (0.088). New homes drive landscaping, finishing, upgrade demand.",
        "leakage_safe": "YES",
        "data_quality": "0% missing after engineering",
        "example_value": "5.95%",
        "alternative_considered": "Used in Role-of-Store growth_score calculation"
    },

    # ═══ COMPETITION (6) ═══
    {
        "no": 30, "group": "F. Competition", "name": "Sister Store Count in TradeArea",
        "type": "Raw", "dtype": "Integer",
        "source": "Direct from CSV",
        "formula": "df['Sister Store Count in TradeArea']",
        "corr_target": -0.029, "importance_pct": 0.04,
        "business_logic": "Number of OTHER Lowe's stores in this store's trade area.",
        "why_included": "SELF-CANNIBALIZATION proxy. 3 Lowe's in one trade area share demand. Critical for true addressable market.",
        "leakage_safe": "YES",
        "data_quality": "0% missing",
        "example_value": "0 (no sister stores)",
        "alternative_considered": "Without this, model would expect identical demographics → identical sales regardless of store density"
    },
    {
        "no": 31, "group": "F. Competition", "name": "total_competitor_ta",
        "type": "Engineered", "dtype": "Integer",
        "source": "Engineered from 17 competitor count columns",
        "formula": "Sum of all 17 'Count in TradeArea' competitor columns",
        "corr_target": -0.035, "importance_pct": 0.06,
        "business_logic": "Total competitor count across all 17 tracked competitors in trade area.",
        "why_included": "COMPOSITE measure. Each individual competitor is weak signal; total pressure is meaningful.",
        "leakage_safe": "YES",
        "data_quality": "0% missing after engineering",
        "example_value": "11 total competitors",
        "alternative_considered": "Used in Role-of-Store risk_score calculation"
    },
    {
        "no": 32, "group": "F. Competition", "name": "Nut Cracker Tools Count in TradeArea",
        "type": "Raw", "dtype": "Integer",
        "source": "Direct from CSV",
        "formula": "df['Nut Cracker Tools Count in TradeArea']",
        "corr_target": -0.024, "importance_pct": 0.04,
        "business_logic": "Count of 'Nut Cracker Tools' (major competitor — Home Depot analog) in trade area.",
        "why_included": "MAJOR head-to-head rival. Mean=3.45, only 10% zeros — highly prevalent and variable.",
        "leakage_safe": "YES",
        "data_quality": "0% missing",
        "example_value": "3",
        "alternative_considered": "Selected over minor competitors due to direct competitive impact"
    },
    {
        "no": 33, "group": "F. Competition", "name": "Wallflowers Depot Count in TradeArea",
        "type": "Raw", "dtype": "Integer",
        "source": "Direct from CSV",
        "formula": "df['Wallflowers Depot Count in TradeArea']",
        "corr_target": -0.006, "importance_pct": 0.41,
        "business_logic": "Count of 'Wallflowers Depot' competitor in trade area.",
        "why_included": "MOST PREVALENT competitor — mean=4.68, only 1% zeros (nearly universal). High importance in tree splits.",
        "leakage_safe": "YES",
        "data_quality": "0% missing",
        "example_value": "5",
        "alternative_considered": "5th most important feature in XGBoost model"
    },
    {
        "no": 34, "group": "F. Competition", "name": "Iggy Pop Hardware Count in TradeArea",
        "type": "Raw", "dtype": "Integer",
        "source": "Direct from CSV",
        "formula": "df['Iggy Pop Hardware Count in TradeArea']",
        "corr_target": -0.043, "importance_pct": 0.04,
        "business_logic": "Count of 'Iggy Pop Hardware' competitor in trade area.",
        "why_included": "HIGHEST CORRELATION among competitors (-0.043). Mean=5.25, 2% zeros — most prevalent.",
        "leakage_safe": "YES",
        "data_quality": "0% missing",
        "example_value": "6",
        "alternative_considered": "Strongest individual competitor signal"
    },
    {
        "no": 35, "group": "F. Competition", "name": "Horn Ok Tools Count in TradeArea",
        "type": "Raw", "dtype": "Integer",
        "source": "Direct from CSV",
        "formula": "df['Horn Ok Tools Count in TradeArea']",
        "corr_target": 0.008, "importance_pct": 0.06,
        "business_logic": "Count of 'Horn Ok Tools' competitor in trade area.",
        "why_included": "HIGH COVERAGE — mean=3.43, 15% zeros — broad market presence.",
        "leakage_safe": "YES",
        "data_quality": "0% missing",
        "example_value": "3",
        "alternative_considered": "Slightly POSITIVE correlation — may indicate 'competitive cluster' areas with overall higher retail activity"
    },
]

# Dropped features
DROPPED = [
    {"name": "Plan Sales USD", "category": "LEAKAGE", "reason": "Correlation with target = 0.986 — it IS the answer. Same-period plan that wouldn't be available at prediction time. Using it would give fake 99% R² but useless in production."},
    {"name": "Invoice Count", "category": "LEAKAGE", "reason": "Correlation = 0.973. Same-period transaction count — directly decomposes sales (Sales = Invoices × Avg Ticket)."},
    {"name": "Avg Ticket", "category": "LEAKAGE", "reason": "Correlation = 0.133. Same-period decomposition of Sales."},
    {"name": "CYE % HH Income 35k-50k", "category": "Multicollinearity", "reason": "Collapsed into income_low_share or income_diy_core to avoid multicollinearity with other income brackets (sum to 100%)."},
    {"name": "CYE % HH Income 100k-150k", "category": "Multicollinearity", "reason": "Middle income bin — captured implicitly by other engineered shares."},
    {"name": "CYE Aggregate Family Household Income", "category": "Redundant", "reason": "PDF marked LOW priority. Aggregate income redundant with Median × Total Households."},
    {"name": "CYE Average Mean Household Income", "category": "Redundant", "reason": "Sensitive to outliers; median is more robust."},
    {"name": "CYE % Housing Unit Built 1970-1979/1980-1989/1990-1999", "category": "Multicollinearity", "reason": "Middle decade bins captured by Median Year Built."},
    {"name": "CYE % Housing Unit Built 2020+", "category": "Sparse", "reason": "Very low share in most areas; merged into housing_new_share concept."},
    {"name": "CYE % House Units Rented", "category": "Redundant", "reason": "Perfect inverse of % Owned (100% redundant)."},
    {"name": "Thin Lizzy Hardware (TA)", "category": "Zero Variance", "reason": "100% zeros across all stores — no information."},
    {"name": "Bobs Supplies (TA)", "category": "Near-Zero Variance", "reason": "99% zeros — no usable signal."},
    {"name": "Lumber Jax / BB King / Eddie Vedder / Mile One / Neil Youngs / JJ Cale (TA)", "category": "Sparse Signal", "reason": "30-77% zeros — mostly absent. Reserved for Phase 2 testing if needed."},
    {"name": "All 3mi/5mi/10mi/0-11mi/11-30mi competitor bands (~80 cols)", "category": "Redundant", "reason": "Per Lowe's methodology, TradeArea captures ~70% of sales. Other distance bands are noisy proxies."},
    {"name": "2010 / 2020 historical demographic snapshots", "category": "Stale", "reason": "CYE (Current Year Estimate) is more relevant. Trend captured by CAGR HH Growth Rate."},
    {"name": "Marital status / Language / Rural % / Education breakdowns", "category": "Weak Signal", "reason": "20+ low-correlation features dropped. Don't drive home improvement spend."},
    {"name": "Real Store ID / Store Name / District", "category": "Identity", "reason": "Captures noise. Meaningful info already in Region/Urbanicity/CBSA."},
    {"name": "CBSA Code / Real CBSA Code", "category": "Identity", "reason": "Identifier — encoded categorical (CBSA Type/Size) captures the relevant signal."},
]

# ──────────────────────────────────────────────────────────────
# Build the workbook
# ──────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# Styles
HEADER_FILL = PatternFill(start_color="003087", end_color="003087", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, color="003087", size=10)
GROUP_FILL_A = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
GROUP_FILL_B = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
GROUP_FILL_C = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
GROUP_FILL_D = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
GROUP_FILL_E = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
GROUP_FILL_F = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

GROUP_COLORS = {
    "A. Calendar": GROUP_FILL_A,
    "B. Engineered Lags": GROUP_FILL_B,
    "C. Store/Market": GROUP_FILL_C,
    "D. Demand/Housing (Raw)": GROUP_FILL_D,
    "E. Demand/Housing (Engineered)": GROUP_FILL_E,
    "F. Competition": GROUP_FILL_F,
}

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# ──────────────────────────────────────────────────────────────
# SHEET 1: Cover/Summary
# ──────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Overview"

ws["A1"] = "Lowe's Store-Level Sales Target Model"
ws["A1"].font = Font(bold=True, color="003087", size=18)
ws["A2"] = "Phase 1 Feature Documentation — All 35 Features"
ws["A2"].font = Font(bold=True, color="005B9F", size=14)
ws["A3"] = "Capstone 8 | Group 8 | IIM Calcutta APAL02"
ws["A3"].font = Font(italic=True, color="64748B", size=11)

ws["A5"] = "DOCUMENT CONTENTS"
ws["A5"].font = HEADER_FONT
ws["A5"].fill = HEADER_FILL

contents = [
    ("Sheet 1: Overview", "Project context and document map"),
    ("Sheet 2: All 35 Features (Master Table)", "Every feature with formula, logic, why, importance"),
    ("Sheet 3: Group A — Calendar (2)", "Year, Fiscal Week"),
    ("Sheet 4: Group B — Engineered Lags (6)", "lag_1/4/13/52, roll_4/13 — the power features"),
    ("Sheet 5: Group C — Store/Market (5)", "Floor size, garden, urbanicity, CBSA"),
    ("Sheet 6: Group D — Demand/Housing Raw (11)", "Demographics, housing, population"),
    ("Sheet 7: Group E — Engineered Demand (5)", "Income mix, housing age shares"),
    ("Sheet 8: Group F — Competition (6)", "Competitors + sister stores"),
    ("Sheet 9: Dropped Features", "What we excluded and exactly why"),
    ("Sheet 10: Feature Engineering Logic", "The 3 principles + multicollinearity handling"),
    ("Sheet 11: Leakage Prevention", "How we kept the model honest"),
    ("Sheet 12: Feature Importance Ranking", "Top 35 features by XGBoost importance"),
]
for i, (sheet, desc) in enumerate(contents, 7):
    ws.cell(row=i, column=1, value=sheet).font = Font(bold=True, color="003087")
    ws.cell(row=i, column=2, value=desc).font = Font(color="334155")

ws["A20"] = "MODEL PERFORMANCE SUMMARY"
ws["A20"].font = HEADER_FONT
ws["A20"].fill = HEADER_FILL

perf_rows = [
    ("Model", "WAPE", "Bias", "RMSE", "R²"),
    ("XGBoost (BEST)", "6.21%", "-2.53%", "$105,749", "0.9822"),
    ("LightGBM", "6.27%", "-2.79%", "$107,411", "0.9816"),
    ("Ridge Regression", "9.13%", "-5.44%", "$130,719", "0.9728"),
    ("Naive Baseline (lag-52)", "8.27%", "+0.33%", "$148,762", "0.9648"),
]
for i, row in enumerate(perf_rows, 22):
    for j, val in enumerate(row, 1):
        cell = ws.cell(row=i, column=j, value=val)
        if i == 22:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        elif "BEST" in str(row[0]):
            cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            cell.font = Font(bold=True)

ws["A28"] = "DATASET STATS"
ws["A28"].font = HEADER_FONT
ws["A28"].fill = HEADER_FILL

stats = [
    ("Total Rows", "269,412"),
    ("Unique Stores", "1,727"),
    ("Time Period", "FY2023 - FY2025 (3 years × 52 weeks)"),
    ("Raw Columns", "194"),
    ("Phase 1 Features Selected", "35 (78% reduction from 159 candidate features)"),
    ("Target Variable", "Actual Sales USD"),
    ("Holdout Strategy", "Rolling time split (85th percentile cutoff = Sep 19, 2025)"),
    ("Holdout Period", "Sep 19, 2025 → Jan 30, 2026 (Q4 FY2025)"),
    ("Holdout Rows", "22,439"),
    ("Stores in Holdout", "1,727 (all)"),
]
for i, (k, v) in enumerate(stats, 30):
    ws.cell(row=i, column=1, value=k).font = Font(bold=True, color="003087")
    ws.cell(row=i, column=2, value=v)

ws.column_dimensions["A"].width = 50
ws.column_dimensions["B"].width = 65
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 15
ws.column_dimensions["E"].width = 15

# ──────────────────────────────────────────────────────────────
# SHEET 2: Master Table — All 35 Features
# ──────────────────────────────────────────────────────────────
ws = wb.create_sheet("All 35 Features")

headers = ["#", "Group", "Feature Name", "Type", "Data Type", "Source",
           "Formula / Calculation", "Correlation w/ Target", "XGBoost Importance %",
           "Business Logic", "Why Included", "Leakage Safe?",
           "Data Quality", "Example Value", "Alternative Considered"]

for j, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=j, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER

for i, f in enumerate(FEATURES, 2):
    row_data = [
        f["no"], f["group"], f["name"], f["type"], f["dtype"], f["source"],
        f["formula"], f["corr_target"], f["importance_pct"],
        f["business_logic"], f["why_included"], f["leakage_safe"],
        f["data_quality"], f["example_value"], f["alternative_considered"]
    ]
    fill = GROUP_COLORS.get(f["group"])
    for j, val in enumerate(row_data, 1):
        cell = ws.cell(row=i, column=j, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
        if j == 9 and isinstance(val, (int, float)):
            cell.number_format = "0.00"

# Column widths
widths = [5, 22, 32, 12, 14, 22, 38, 12, 12, 50, 50, 12, 28, 20, 50]
for j, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.row_dimensions[1].height = 38
for i in range(2, len(FEATURES) + 2):
    ws.row_dimensions[i].height = 90
ws.freeze_panes = "D2"

# ──────────────────────────────────────────────────────────────
# SHEETS 3-8: Per-Group Detailed Sheets
# ──────────────────────────────────────────────────────────────
groups = [
    ("Calendar (2)", "A. Calendar", "Captures macro time trends and weekly seasonality. These features tell the model WHEN the prediction is happening — without them, the model can't tell February from May."),
    ("Engineered Lags (6)", "B. Engineered Lags", "THE POWER FEATURES — 94% of total feature importance. These use a store's PAST sales to predict its FUTURE sales. All apply shift(1) minimum to prevent leakage. Why 6 lags? Each captures a different time horizon: weekly (lag_1), monthly (lag_4), quarterly (lag_13), annual (lag_52), and smoothed trends (roll_4, roll_13)."),
    ("Store-Market (5)", "C. Store/Market", "Physical and locational attributes of each store. These are STATIC features (don't change weekly) but provide essential context — they normalize the lag features. A small store with $1M sales is overperforming; a large store with same number is underperforming."),
    ("Demand-Housing Raw (11)", "D. Demand/Housing (Raw)", "Demographics and housing characteristics from the trade area. These capture WHO lives near each store and WHAT their housing needs are. Sources: U.S. Census Bureau Current Year Estimates (CYE)."),
    ("Demand-Housing Engineered (5)", "E. Demand/Housing (Engineered)", "Engineered combinations of raw demographic features. Each replaces 2-3 raw columns to AVOID MULTICOLLINEARITY (the original income brackets sum to 100%, housing decades sum to 100%). The engineered features capture strategic segments without destabilizing the model."),
    ("Competition (6)", "F. Competition", "Competitive landscape in each store's trade area. Selected from 17 raw competitors using prevalence (mean count) and zero-rate. Major competitors get individual columns; total composite captures aggregate pressure."),
]

for sheet_name, group_key, description in groups:
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = f"GROUP {group_key}"
    ws["A1"].font = Font(bold=True, color="003087", size=16)
    ws["A2"] = description
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A2"].font = Font(color="334155", size=11)
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 60

    headers = ["#", "Feature", "Formula", "Corr", "Importance %",
               "Business Logic", "Why Included", "Example"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=j, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    row = 5
    for f in FEATURES:
        if f["group"] == group_key:
            ws.cell(row=row, column=1, value=f["no"])
            ws.cell(row=row, column=2, value=f["name"])
            ws.cell(row=row, column=3, value=f["formula"])
            ws.cell(row=row, column=4, value=f["corr_target"])
            ws.cell(row=row, column=5, value=f["importance_pct"])
            ws.cell(row=row, column=6, value=f["business_logic"])
            ws.cell(row=row, column=7, value=f["why_included"])
            ws.cell(row=row, column=8, value=f["example_value"])
            for j in range(1, 9):
                cell = ws.cell(row=row, column=j)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = THIN_BORDER
                cell.fill = GROUP_COLORS[group_key]
            ws.row_dimensions[row].height = 90
            row += 1

    widths = [5, 28, 38, 10, 14, 50, 50, 22]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

# ──────────────────────────────────────────────────────────────
# SHEET 9: Dropped Features
# ──────────────────────────────────────────────────────────────
ws = wb.create_sheet("Dropped Features")
ws["A1"] = "Features We DROPPED from Phase 1"
ws["A1"].font = Font(bold=True, color="C62828", size=16)
ws["A2"] = "Out of ~159 candidate features, we kept only 35. Here's exactly what was dropped and why."
ws["A2"].alignment = Alignment(wrap_text=True)
ws.merge_cells("A2:C2")

headers = ["Feature/Group", "Drop Category", "Reason"]
for j, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=j, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = THIN_BORDER

for i, d in enumerate(DROPPED, 5):
    cells = [
        ws.cell(row=i, column=1, value=d["name"]),
        ws.cell(row=i, column=2, value=d["category"]),
        ws.cell(row=i, column=3, value=d["reason"]),
    ]
    for cell in cells:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = THIN_BORDER
        if d["category"] == "LEAKAGE":
            cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        elif d["category"] in ("Multicollinearity", "Redundant"):
            cell.fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
    ws.row_dimensions[i].height = 60

ws.column_dimensions["A"].width = 60
ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 90

# ──────────────────────────────────────────────────────────────
# SHEET 10: Engineering Logic
# ──────────────────────────────────────────────────────────────
ws = wb.create_sheet("Engineering Logic")
ws["A1"] = "Feature Engineering Philosophy & Logic"
ws["A1"].font = Font(bold=True, color="003087", size=16)

content = [
    ("THE 3 RULES", "All feature engineering follows these principles", True),
    ("", "", False),
    ("Rule 1: NO LEAKAGE", "Never use information that wouldn't be available at prediction time. Every lag feature uses shift(1) minimum so the model can't 'cheat' by seeing the answer.", False),
    ("Rule 2: REDUCE REDUNDANCY", "When raw features sum to 100% (income brackets, housing decades), engineer 2-3 strategic combinations. Multicollinearity destabilizes models — it doesn't add information, it confuses the model.", False),
    ("Rule 3: BUSINESS MEANING", "Every engineered feature must map to a concept a planner would recognize: 'core DIY customer share', 'competitive pressure', 'annual seasonality'. If you can't explain it in one sentence, it shouldn't be in the model.", False),
    ("", "", False),
    ("MULTICOLLINEARITY HANDLING", "How we collapsed redundant features", True),
    ("", "", False),
    ("Income brackets (8 → 3)", "Original: 8 income share columns sum to ~100%. Solution: collapsed into income_low_share (<$35K), income_diy_core ($50K-$100K), income_affluent ($150K+). Median income still kept as central tendency.", False),
    ("Housing decades (9 → 2)", "Original: 9 decade columns sum to ~100%. Solution: housing_old_share (pre-1969), housing_new_share (post-2000). Median Year Built kept for central tendency. Middle decades captured implicitly.", False),
    ("Competitor columns (98 → 6)", "Original: 17 competitors × 5-6 distance bands = 98 columns. Solution: kept only TradeArea (Lowe's gold standard), selected 4 most prevalent competitors + Sister Stores + total_competitor_ta composite.", False),
    ("", "", False),
    ("LAG FEATURE DESIGN", "Why we use shift(1) before any rolling window", True),
    ("", "", False),
    ("WRONG (leaks)", "df.groupby('Store ID')['Sales'].rolling(4).mean()  ← Includes current week!", False),
    ("RIGHT (safe)", "df.groupby('Store ID')['Sales'].transform(lambda x: x.shift(1).rolling(4).mean())", False),
    ("Why?", "If we predict week t and the rolling window includes week t, we're using the answer to predict the answer. shift(1) ensures the rolling window only sees weeks t-4 through t-1.", False),
    ("", "", False),
    ("CATEGORICAL ENCODING", "How we converted text → numbers for trees", True),
    ("", "", False),
    ("LabelEncoder approach", "Tree models split on numeric thresholds. We assigned integer codes: Metropolis=6, Large City=5, ..., Remote=0. Trees don't assume linear ordering — they find optimal splits regardless.", False),
    ("Why not One-Hot?", "One-hot creates many sparse columns that hurt tree models. LabelEncoder is preferred for ordinal-ish categorical features in tree-based models.", False),
]

for i, (k, v, header) in enumerate(content, 3):
    if header:
        cell = ws.cell(row=i, column=1, value=k)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
        cell2 = ws.cell(row=i + 1, column=1, value=v) if v else None
    else:
        if k:
            cell = ws.cell(row=i, column=1, value=k)
            cell.font = Font(bold=True, color="003087")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell = ws.cell(row=i, column=2, value=v)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if v:
            ws.row_dimensions[i].height = 40

ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 100

# ──────────────────────────────────────────────────────────────
# SHEET 11: Leakage Prevention
# ──────────────────────────────────────────────────────────────
ws = wb.create_sheet("Leakage Prevention")
ws["A1"] = "Data Leakage Prevention — The Critical Guard"
ws["A1"].font = Font(bold=True, color="C62828", size=16)
ws["A2"] = "Three columns are NEVER used as inputs. Using them would give fake high accuracy but make the model useless in production."
ws["A2"].alignment = Alignment(wrap_text=True)
ws.merge_cells("A2:D2")
ws.row_dimensions[2].height = 35

headers = ["Forbidden Feature", "Correlation w/ Target", "Why It's Leakage", "What Would Happen If Used"]
for j, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=j, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = THIN_BORDER

leak_data = [
    ("Plan Sales USD", "0.986",
     "It IS the target — the planned sales for this very same week. The 'plan' is essentially what we're trying to PREDICT (or improve upon).",
     "Model would get fake 99% R² but be USELESS in production. When predicting future weeks, you don't HAVE next week's plan — that's what the model is supposed to generate!"),
    ("Invoice Count", "0.972",
     "Same-period transaction count. Sales = Invoices × Avg Ticket — using Invoice Count gives away half the answer.",
     "Model would 'predict' sales by basically reading off transaction counts. Doesn't help with forward-looking planning."),
    ("Avg Ticket", "0.133",
     "Same-period decomposition of Sales. While correlation is lower, it still represents 'inside knowledge' of the same week.",
     "Model would be partly cheating — using mid-week or end-week observations to predict the same week."),
]

for i, row in enumerate(leak_data, 5):
    for j, val in enumerate(row, 1):
        cell = ws.cell(row=i, column=j, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = THIN_BORDER
        cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    ws.row_dimensions[i].height = 90

ws["A10"] = "GUARD CODE IN feature_engineering.py"
ws["A10"].font = HEADER_FONT
ws["A10"].fill = HEADER_FILL
ws.merge_cells("A10:D10")

ws["A11"] = """LEAKAGE_BLACKLIST = ['Plan Sales USD', 'Invoice Count', 'Avg Ticket']

# Programmatic check that runs every time features are built:
for col in LEAKAGE_BLACKLIST:
    if col in PHASE1_FEATURES:
        raise ValueError(f"LEAKAGE DETECTED: {col} is in PHASE1_FEATURES! Remove it immediately.")"""
ws["A11"].font = Font(name="Courier New", size=10)
ws["A11"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("A11:D11")
ws.row_dimensions[11].height = 100

ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 50
ws.column_dimensions["D"].width = 50

# ──────────────────────────────────────────────────────────────
# SHEET 12: Feature Importance Ranking
# ──────────────────────────────────────────────────────────────
ws = wb.create_sheet("Importance Ranking")
ws["A1"] = "Feature Importance — Ranked by XGBoost"
ws["A1"].font = Font(bold=True, color="003087", size=16)
ws["A2"] = "How much each feature contributed to the model's decisions, expressed as a percentage."
ws["A2"].alignment = Alignment(wrap_text=True)
ws.merge_cells("A2:D2")

headers = ["Rank", "Feature", "Group", "Importance %"]
for j, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=j, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    cell.border = THIN_BORDER

ranked = sorted(FEATURES, key=lambda f: f["importance_pct"], reverse=True)
for i, f in enumerate(ranked, 5):
    ws.cell(row=i, column=1, value=i - 4)
    ws.cell(row=i, column=2, value=f["name"])
    ws.cell(row=i, column=3, value=f["group"])
    ws.cell(row=i, column=4, value=f["importance_pct"]).number_format = "0.00"

    fill = GROUP_COLORS.get(f["group"])
    for j in range(1, 5):
        cell = ws.cell(row=i, column=j)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(vertical="center")

    if i - 4 <= 3:
        for j in range(1, 5):
            ws.cell(row=i, column=j).font = Font(bold=True, size=11)

ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 35
ws.column_dimensions["C"].width = 30
ws.column_dimensions["D"].width = 15

ws["A42"] = "INSIGHT"
ws["A42"].font = HEADER_FONT
ws["A42"].fill = HEADER_FILL
ws["A43"] = ("The top 2 features (lag_1 + roll_4) account for 93.6% of model importance. "
             "This tells us the model's primary 'reasoning' is: 'last week's sales + recent trend = next week's sales'. "
             "Demographics and competition matter less because — for 1-week-ahead prediction — recent store performance "
             "already encodes most of those slow-moving market effects. Demographics would matter MORE for predicting a NEW store "
             "(no lag history) or for long-horizon forecasts (multiple years ahead).")
ws["A43"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("A43:D43")
ws.row_dimensions[43].height = 100

# Save
output_path = "outputs/Phase1_Feature_Documentation.xlsx"
wb.save(output_path)
print(f"✓ Saved: {output_path}")
print(f"  Sheets: {len(wb.sheetnames)}")
for name in wb.sheetnames:
    print(f"    - {name}")

# Also copy to Desktop for easy access
import shutil
desktop = os.path.expanduser("~/Desktop/Phase1_Feature_Documentation.xlsx")
shutil.copy(output_path, desktop)
print(f"\n✓ Also saved to Desktop: {desktop}")
print(f"  File size: {os.path.getsize(desktop) / 1024:.1f} KB")
